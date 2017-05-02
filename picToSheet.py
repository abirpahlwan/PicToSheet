import openpyxl as excel
import cv2
import numpy as np

fileName = "DSC_3267-001"

img = cv2.imread(fileName + '.jpg')

height, width, _ = img.shape
if height > 1024 or width > 1024:
	img = cv2.resize(img, (1024, int(1024*height/width))) if width >= height else cv2.resize(img, (int(1024*width/height), 1024))

wBook = excel.Workbook()
wSheet = wBook.active
wSheet.title = fileName

height, width, _ = img.shape
for x in range(0,width):
	for y in range(0,height):
		b, g, r = img[y, x]

		hexR = str("0" + hex(r)[2:])[-2:]
		hexG = str("0" + hex(g)[2:])[-2:]
		hexB = str("0" + hex(b)[2:])[-2:]

		row1 = wSheet.cell(row=(y*3)+1, column=x+1, value="")
		strR = hexR + "00" + "00"
		row2 = wSheet.cell(row=(y*3)+2, column=x+1, value="")
		strG = "00" + hexG + "00"
		row3 = wSheet.cell(row=(y*3)+3, column=x+1, value="")
		strB = "00" + "00" + hexB

		row1.fill = excel.styles.PatternFill(start_color=strR, end_color=strR, fill_type='solid')
		row2.fill = excel.styles.PatternFill(start_color=strG, end_color=strG, fill_type='solid')
		row3.fill = excel.styles.PatternFill(start_color=strB, end_color=strB, fill_type='solid')
		pass
	print((x+1)*100/width, "%")
	pass

print("Wait a little")
wBook.save(fileName + '.xlsx')
print("Export Complete")

# cv2.imshow(fileName, img)
# cv2.waitKey(0)
# cv2.destroyAllWindows()